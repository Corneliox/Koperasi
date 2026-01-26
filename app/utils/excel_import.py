"""
Advanced Excel Import Engine for Inventory
REFACTORED: Added Smart Import with Sheet Selection and Data Preview

Expected Format:
Row 1 (A1-K1): Title "STOK BARANG BULAN {Month} {Year}"
Row 4 (Headers): NO | NAME | START STOCK | IN | OUT | END STOCK | BUY PRICE | SELL PRICE | ITEM PROFIT | TOTAL PROFIT | ASSET

Business Logic (STRICT):
- Returns: Subtract from IN column, NOT add to OUT
- End Stock = Start Stock + IN - OUT
- Item Profit = Sell Price - Buy Price
- Total Profit = Item Profit * OUT
- Asset = Sell Price * End Stock
"""
import os
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from app.database.connection import get_connection, log_activity


def get_workbook_sheets(filepath: str) -> list:
    """
    Get list of worksheet names from Excel file
    :param filepath: Path to Excel file
    :return: List of sheet names
    """
    try:
        wb = load_workbook(filepath, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        return []


def preview_excel_data(filepath: str, sheet_name: str = None, max_rows: int = 20) -> dict:
    """
    Preview data from Excel file before import
    
    :param filepath: Path to Excel file
    :param sheet_name: Specific sheet to preview (None = active sheet)
    :param max_rows: Maximum rows to preview
    :return: Dict with headers and data preview
    """
    try:
        # Load with specific sheet if provided
        if sheet_name:
            df = pd.read_excel(filepath, sheet_name=sheet_name, skiprows=3, 
                              engine='openpyxl', nrows=max_rows)
        else:
            df = pd.read_excel(filepath, skiprows=3, engine='openpyxl', nrows=max_rows)
        
        if df is None or df.empty:
            # Try reading without skipping
            if sheet_name:
                df = pd.read_excel(filepath, sheet_name=sheet_name, 
                                  engine='openpyxl', nrows=max_rows)
            else:
                df = pd.read_excel(filepath, engine='openpyxl', nrows=max_rows)
        
        if df is None or df.empty:
            return {"success": False, "message": "File kosong atau tidak dapat dibaca"}
        
        # Normalize column names
        df.columns = [str(col).strip().upper() for col in df.columns]
        
        # Convert to list of dicts for preview
        headers = list(df.columns)
        data = df.fillna('').astype(str).values.tolist()
        
        # Count total rows
        if sheet_name:
            df_full = pd.read_excel(filepath, sheet_name=sheet_name, 
                                   skiprows=3, engine='openpyxl')
        else:
            df_full = pd.read_excel(filepath, skiprows=3, engine='openpyxl')
        
        total_rows = len(df_full) if df_full is not None else 0
        
        return {
            "success": True,
            "headers": headers,
            "data": data,
            "preview_count": len(data),
            "total_rows": total_rows
        }
        
    except Exception as e:
        return {"success": False, "message": str(e)}


def import_inventory_from_excel(filepath: str, category_context: str, 
                                 current_user: str = "admin") -> dict:
    """
    Import inventory from Excel file
    
    :param filepath: Path to Excel file
    :param category_context: 'SEMBAKO' or 'TAKTIKAL'
    :param current_user: User performing import
    :return: Result dict with success status and counts
    """
    try:
        # Load workbook to check format
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        
        # Check for title row (Row 1)
        title_cell = ws['A1'].value
        month_year = ""
        if title_cell and "STOK BARANG BULAN" in str(title_cell).upper():
            month_year = str(title_cell).replace("STOK BARANG BULAN", "").strip()
        
        wb.close()
        
        # Read data starting from row 4 (headers) or try to auto-detect
        df = None
        
        # Try reading with header at row 3 (0-indexed: skiprows=3)
        try:
            df = pd.read_excel(filepath, skiprows=3, engine='openpyxl')
        except:
            # Fallback: try reading without skipping
            df = pd.read_excel(filepath, engine='openpyxl')
        
        if df is None or df.empty:
            return {"success": False, "message": "File Excel kosong atau format tidak valid"}
        
        # Normalize column names
        df.columns = [str(col).strip().upper() for col in df.columns]
        
        # Map possible column names
        column_mapping = {
            'NO': ['NO', 'NOMOR', '#'],
            'NAME': ['NAME', 'NAMA', 'NAMA BARANG', 'ITEM'],
            'START_STOCK': ['START STOCK', 'STOK AWAL', 'STOCK AWAL', 'AWAL'],
            'IN': ['IN', 'MASUK', 'INCOMING', 'BARANG MASUK'],
            'OUT': ['OUT', 'KELUAR', 'OUTGOING', 'BARANG KELUAR', 'TERJUAL'],
            'END_STOCK': ['END STOCK', 'STOK AKHIR', 'STOCK AKHIR', 'AKHIR'],
            'BUY_PRICE': ['BUY PRICE', 'HARGA BELI', 'HARGA MODAL', 'MODAL'],
            'SELL_PRICE': ['SELL PRICE', 'HARGA JUAL', 'HARGA'],
            'STATUS': ['STATUS', 'KEPEMILIKAN', 'TIPE']
        }
        
        # Find matching columns
        def find_column(possible_names):
            for name in possible_names:
                if name in df.columns:
                    return name
            return None
        
        col_name = find_column(column_mapping['NAME'])
        col_start = find_column(column_mapping['START_STOCK'])
        col_in = find_column(column_mapping['IN'])
        col_out = find_column(column_mapping['OUT'])
        col_end = find_column(column_mapping['END_STOCK'])
        col_buy = find_column(column_mapping['BUY_PRICE'])
        col_sell = find_column(column_mapping['SELL_PRICE'])
        col_status = find_column(column_mapping['STATUS'])
        
        if not col_name:
            return {"success": False, "message": "Kolom 'NAME' atau 'NAMA BARANG' tidak ditemukan"}
        
        # Process data
        conn = get_connection()
        cursor = conn.cursor()
        
        added = 0
        updated = 0
        errors = []
        
        for idx, row in df.iterrows():
            try:
                name = str(row.get(col_name, '')).strip()
                if not name or name == 'nan' or pd.isna(row.get(col_name)):
                    continue
                
                # Get values with defaults
                start_stock = int(row.get(col_start, 0) or 0) if col_start else 0
                stock_in = int(row.get(col_in, 0) or 0) if col_in else 0
                stock_out = int(row.get(col_out, 0) or 0) if col_out else 0
                
                # Calculate end stock using business logic
                # End Stock = Start Stock + IN - OUT (returns are subtracted from IN)
                if col_end:
                    end_stock = int(row.get(col_end, 0) or 0)
                else:
                    end_stock = start_stock + stock_in - stock_out
                
                buy_price = float(row.get(col_buy, 0) or 0) if col_buy else 0
                sell_price = float(row.get(col_sell, 0) or 0) if col_sell else 0
                
                status = str(row.get(col_status, 'Koperasi') or 'Koperasi').strip() if col_status else 'Koperasi'
                if status not in ['Koperasi', 'Konsinyasi']:
                    status = 'Koperasi'
                
                # Calculate business values
                item_profit = sell_price - buy_price
                total_profit = item_profit * stock_out
                asset = sell_price * end_stock
                
                # Check if item exists
                cursor.execute(
                    "SELECT id, stock FROM warehouse WHERE name = ? AND category_type = ?",
                    (name, category_context)
                )
                existing = cursor.fetchone()
                
                if existing:
                    # Update existing item
                    old_stock = existing['stock']
                    cursor.execute(
                        """UPDATE warehouse 
                           SET stock=?, buy_price=?, sell_price=?, status=?, 
                               updated_at=CURRENT_TIMESTAMP
                           WHERE id=?""",
                        (end_stock, buy_price, sell_price, status, existing['id'])
                    )
                    
                    # Create mutation if stock changed
                    stock_diff = end_stock - old_stock
                    if stock_diff != 0:
                        mutation_type = 'IN' if stock_diff > 0 else 'OUT'
                        cursor.execute(
                            """INSERT INTO warehouse_mutation (item_id, type, qty, description)
                               VALUES (?, ?, ?, ?)""",
                            (existing['id'], mutation_type, abs(stock_diff), 
                             f"Import Excel: {month_year}")
                        )
                    
                    updated += 1
                else:
                    # Insert new item
                    cursor.execute(
                        """INSERT INTO warehouse 
                           (name, category_type, stock, buy_price, sell_price, status, description)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (name, category_context, end_stock, buy_price, sell_price, status,
                         f"Import: {month_year}")
                    )
                    item_id = cursor.lastrowid
                    
                    # Create initial mutation
                    if end_stock > 0:
                        cursor.execute(
                            """INSERT INTO warehouse_mutation (item_id, type, qty, description)
                               VALUES (?, 'IN', ?, ?)""",
                            (item_id, end_stock, f"Import Excel: {month_year}")
                        )
                    
                    added += 1
                    
            except Exception as e:
                errors.append(f"Baris {idx + 5}: {str(e)}")
        
        conn.commit()
        conn.close()
        
        # Log activity
        log_activity(
            current_user,
            "IMPORT_EXCEL",
            f"Import inventaris dari Excel: {added} ditambah, {updated} diupdate"
        )
        
        result = {
            "success": True,
            "message": "Import berhasil",
            "total_items": added + updated,
            "added": added,
            "updated": updated,
            "month_year": month_year
        }
        
        if errors:
            result["warnings"] = errors[:10]  # Limit to first 10 errors
        
        return result
        
    except Exception as e:
        return {"success": False, "message": str(e)}


def export_monthly_stock_report(category_context: str, month: int, year: int,
                                output_dir: str = None) -> str:
    """
    Export monthly stock report in the expected format
    
    Format:
    Row 1: "STOK BARANG BULAN {Month} {Year}"
    Row 4: Headers
    
    Business Logic:
    - End Stock = Start Stock + IN - OUT
    - Item Profit = Sell Price - Buy Price
    - Total Profit = Item Profit * OUT
    - Asset = Sell Price * End Stock
    """
    import calendar
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    
    if output_dir is None:
        output_dir = os.path.join(os.path.expanduser("~"), "Documents", "Koperasi_Export")
    
    os.makedirs(output_dir, exist_ok=True)
    
    conn = get_connection()
    cursor = conn.cursor()
    
    # Get all items for this category
    cursor.execute(
        "SELECT * FROM warehouse WHERE category_type = ? ORDER BY name",
        (category_context,)
    )
    items = [dict(row) for row in cursor.fetchall()]
    
    # Get mutations for the month
    month_start = f"{year}-{month:02d}-01"
    month_end = f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]}"
    
    # Prepare data
    data = []
    total_profit = 0
    total_asset = 0
    
    for idx, item in enumerate(items):
        # Get mutations for this item in this month
        cursor.execute(
            """SELECT type, SUM(qty) as total_qty 
               FROM warehouse_mutation 
               WHERE item_id = ? AND DATE(date) BETWEEN ? AND ?
               GROUP BY type""",
            (item['id'], month_start, month_end)
        )
        mutations = {row['type']: row['total_qty'] for row in cursor.fetchall()}
        
        stock_in = mutations.get('IN', 0)
        stock_out = mutations.get('OUT', 0)
        returns = mutations.get('RETURN', 0)
        
        # Business logic: Returns subtract from IN
        effective_in = stock_in - returns
        
        # Current stock is end stock
        end_stock = item['stock']
        
        # Calculate start stock (reverse engineering)
        start_stock = end_stock - effective_in + stock_out
        
        buy_price = item['buy_price']
        sell_price = item['sell_price']
        
        item_profit = sell_price - buy_price
        row_total_profit = item_profit * stock_out
        asset = sell_price * end_stock
        
        total_profit += row_total_profit
        total_asset += asset
        
        data.append({
            'NO': idx + 1,
            'NAME': item['name'],
            'START STOCK': start_stock,
            'IN': effective_in,
            'OUT': stock_out,
            'END STOCK': end_stock,
            'BUY PRICE': buy_price,
            'SELL PRICE': sell_price,
            'ITEM PROFIT': item_profit,
            'TOTAL PROFIT': row_total_profit,
            'ASSET': asset
        })
    
    conn.close()
    
    # Create Excel file
    month_name = calendar.month_name[month]
    df = pd.DataFrame(data)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Stok_Barang_{category_context}_{month_name}_{year}_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Write data starting at row 4
        df.to_excel(writer, sheet_name='Stok Barang', index=False, startrow=3)
        
        workbook = writer.book
        worksheet = writer.sheets['Stok Barang']
        
        # Add title at row 1
        worksheet['A1'] = f"STOK BARANG BULAN {month_name.upper()} {year}"
        worksheet['A1'].font = Font(size=16, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center')
        worksheet.merge_cells('A1:K1')
        
        # Add category info at row 2
        worksheet['A2'] = f"Kategori: {category_context}"
        worksheet['A2'].font = Font(size=12)
        
        # Style headers at row 4
        header_fill = PatternFill(start_color='2B579A', end_color='2B579A', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, 12):
            cell = worksheet.cell(row=4, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')
        
        # Format data cells
        for row in range(5, len(data) + 5):
            for col in range(1, 12):
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                # Format numbers
                if col >= 3:  # Number columns
                    cell.number_format = '#,##0'
        
        # Add totals row
        total_row = len(data) + 5
        worksheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        worksheet.cell(row=total_row, column=10, value=total_profit).number_format = '#,##0'
        worksheet.cell(row=total_row, column=11, value=total_asset).number_format = '#,##0'
        
        # Adjust column widths
        column_widths = [5, 40, 12, 10, 10, 12, 15, 15, 15, 15, 18]
        for i, width in enumerate(column_widths):
            worksheet.column_dimensions[chr(65 + i)].width = width
    
    return filepath
